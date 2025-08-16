import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import re
import os
from tkinter import ttk
import json
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class GDTExtractor:
    def __init__(self):
        self.output_table = None
        self.dark_mode = False
        self.logo_photo = None
        self.right_photo = None
        
        # GD&T symbol mapping - expanded to include dimensional symbols
        self.gdnt_symbols = {
            "Straightness": "─",
            "Flatness": "☐", 
            "Circularity": "○",
            "Cylindricity": "⌀",
            "Diameter": "⌀",
            "Length": "↔",
            "Linear Distance": "↔"
        }
        
    def calculate_tolerance_limits(self, tolerance_type, tolerance_value, nominal_value=0):
        """Calculate upper and lower limits based on tolerance type and value"""
        try:
            tol_val = float(tolerance_value) if tolerance_value != "N/A" else 0
            nominal = float(nominal_value)
            
            # For geometric tolerances, limits are typically ±tolerance/2 from nominal
            # For form tolerances (flatness, circularity, etc.), it's usually 0 to +tolerance
            tolerance_type_lower = tolerance_type.lower()
            
            if tolerance_type_lower in ["flatness", "straightness", "circularity", "cylindricity"]:
                # Form tolerances: deviation from perfect form (0 to +tolerance)
                upper_limit = nominal + tol_val
                lower_limit = nominal  # Perfect form is the lower limit
            elif tolerance_type_lower in ["position", "concentricity", "symmetry"]:
                # Location tolerances: ±tolerance/2
                upper_limit = nominal + (tol_val / 2)
                lower_limit = nominal - (tol_val / 2)
            elif tolerance_type_lower == "datum":
                # Datums are reference points - no tolerance limits
                return "REF", "REF"
            else:
                # Default case: ±tolerance/2
                upper_limit = nominal + (tol_val / 2)
                lower_limit = nominal - (tol_val / 2)
            
            # Format to reasonable decimal places
            upper_limit = round(upper_limit, 4)
            lower_limit = round(lower_limit, 4)
            
            return str(upper_limit), str(lower_limit)
            
        except (ValueError, TypeError):
            return "N/A", "N/A"

    def format_tolerance_value(self, value):
        """Format tolerance value with ± symbol"""
        if value == "N/A" or not value:
            return "N/A"
        try:
            # Convert to float to validate it's a number
            float_val = float(value)
            # Format with ± symbol
            return f"±{value}"
        except (ValueError, TypeError):
            return str(value)
        
    def determine_plane_position(self, feature_name, datum_letter=None, plane_data=None):
        """Enhanced function to determine if a plane datum is on top face or bottom face"""
        if not feature_name:
            return "surface"
            
        fname = str(feature_name).lower()
        
        # Direct keyword detection for position
        if any(keyword in fname for keyword in ["top", "upper", "above"]):
            return "top face"
        elif any(keyword in fname for keyword in ["bottom", "lower", "below", "base"]):
            return "bottom face"
        
        # Enhanced datum letter logic - based on common CAD conventions
        if datum_letter:
            datum_upper = datum_letter.upper()
            # In many CAD systems, datum A is often the primary reference (base/bottom)
            # and subsequent datums (B, C, D, etc.) can be on different surfaces
            if datum_upper == 'A':
                # Check if it's explicitly a top surface first
                if any(keyword in fname for keyword in ["top", "upper"]):
                    return "top face"
                else:
                    return "bottom face"  # A is typically base/bottom
            elif datum_upper in ['B', 'C']:
                return "cylindrical side"  # B and C often on cylindrical surfaces
            elif datum_upper == 'D':
                # D is often the secondary planar surface (opposite of A)
                return "top face"
            elif datum_upper in ['E', 'F', 'G', 'H']:
                # Additional datums might alternate
                return "top face" if datum_upper in ['E', 'G'] else "bottom face"
        
        # Plane numbering logic (less reliable)
        elif "plane1" in fname:
            return "bottom face"  # Often plane1 is the base
        elif "plane2" in fname:
            return "top face"   # Often plane2 is the top
        
        # Check for numerical patterns that might indicate position
        plane_numbers = re.findall(r'plane(\d+)', fname)
        if plane_numbers:
            plane_num = int(plane_numbers[0])
            if plane_num == 1:
                return "bottom face"  # Reversed assumption
            elif plane_num >= 2:
                return "top face"
        
        # If it contains "plane" but no clear position indicator, try to infer
        if "plane" in fname:
            # Check if there are any Z-axis indicators or position clues
            if any(indicator in fname for indicator in ["+z", "positive", "high"]):
                return "top face"
            elif any(indicator in fname for indicator in ["-z", "negative", "low"]):
                return "bottom face"
            else:
                # Default assumption: return as generic plane
                return "planar surface"
        
        return str(feature_name)

    def clean_feature_name(self, feature_name):
        """Clean and parse feature names"""
        if not feature_name:
            return ""
        
        fname = str(feature_name).lower()
        fname = re.sub(r'datum\d+@', '', fname)  # Remove datum references
        
        feature_types = ["torus", "plane", "boss", "cylinder", "cone", "sphere"]
        for ftype in feature_types:
            if ftype in fname:
                return ftype
        
        return feature_name

    def get_surface_type(self, feature_name, datum_letter=None):
        """Enhanced helper to map feature name to surface type (for Location column)"""
        if not feature_name:
            return "surface"
            
        fname = str(feature_name).lower()
        
        # Check for torus patterns first
        if "torus" in fname:
            return "torus side"
        # Check for cleaned patterns
        cleaned_name = self.clean_feature_name(feature_name)
        if cleaned_name == "torus":
            return "torus side"
        
        # Enhanced plane detection for datums
        if "plane" in fname:
            return self.determine_plane_position(feature_name, datum_letter)
            
        if "cone" in fname or "conical" in fname:
            return "conical side of the part"
        elif "boss1" in fname or "cylindrical" in fname or "side" in fname:
            return "cylindrical side"
        else:
            return str(feature_name)

    def get_likely_location(self, label, feature_name):
        """Enhanced helper to map feature name and type to Surface (for Surface column)"""
        if not feature_name:
            return "surface"
            
        fname = str(feature_name).lower()
        
        # Check for torus patterns first
        if "torus" in fname:
            return "torus side"
        # Check for cleaned patterns
        cleaned_name = self.clean_feature_name(feature_name)
        if cleaned_name == "torus":
            return "torus side"
        
        # Enhanced plane detection for datums
        if "plane" in fname:
            return self.determine_plane_position(feature_name)
            
        if "cone" in fname or "conical" in fname:
            return "conical side of the part"
        elif "boss1" in fname or "cylindrical" in fname or "side" in fname:
            return "curved side of the cylinder"
        elif "face" in fname:
            return "planar face"
        else:
            return str(feature_name)

    def find_datum_for_dimensional_tolerance(self, feature_name, datum_results, datum_letter_to_feature):
        """NEW: Find appropriate datum for dimensional tolerances based on feature name and type"""
        if not feature_name:
            return ""
        
        fname = str(feature_name).lower()
        
        # Method 1: Direct datum reference extraction from feature name
        # Look for patterns like "Datum29@Boss1(A)" in feature name
        datum_match = re.search(r'datum\d+@[^(]*\(([A-Z])\)', fname)
        if datum_match:
            return datum_match.group(1)
        
        # Method 2: Look for datum letter in parentheses
        paren_match = re.search(r'\(([A-Z])\)', fname)
        if paren_match:
            candidate_datum = paren_match.group(1)
            if candidate_datum in datum_results:
                return candidate_datum
        
        # Method 3: Match based on feature type compatibility
        if "boss" in fname or "cylinder" in fname:
            # For cylindrical features, look for cylindrical datums
            for datum_letter, location in datum_results.items():
                if "cylindrical" in location.lower():
                    return datum_letter
            # If no cylindrical datum found, use the first available
            if datum_results:
                return list(datum_results.keys())[0]
        
        elif "plane" in fname:
            # For planar features, look for planar datums
            plane_num_match = re.search(r'plane(\d+)', fname)
            if plane_num_match:
                plane_num = int(plane_num_match.group(1))
                # Try to match plane number with datum sequence
                datum_letters = sorted(datum_results.keys())
                if plane_num <= len(datum_letters):
                    return datum_letters[plane_num - 1]
            
            # Look for planar datums
            for datum_letter, location in datum_results.items():
                if any(keyword in location.lower() for keyword in ["face", "plane"]):
                    return datum_letter
        
        # Method 4: Match by feature name similarity
        for datum_letter, datum_feature_name in datum_letter_to_feature.items():
            datum_feature_lower = datum_feature_name.lower()
            
            # Extract core feature identifiers
            if "boss" in fname and "boss" in datum_feature_lower:
                return datum_letter
            elif "plane1" in fname and "plane1" in datum_feature_lower:
                return datum_letter
            elif "plane2" in fname and "plane2" in datum_feature_lower:
                return datum_letter
            elif "plane" in fname and "plane" in datum_feature_lower:
                return datum_letter
        
        # Method 5: Default assignment based on common CAD conventions
        if datum_results:
            # For diameter measurements, typically use the primary datum (A) or cylindrical datum
            if any(keyword in fname for keyword in ["boss", "cylinder", "diameter"]):
                # Look for datum A first (often primary)
                if 'A' in datum_results:
                    return 'A'
                # Then look for any cylindrical datum
                for datum_letter, location in datum_results.items():
                    if "cylindrical" in location.lower():
                        return datum_letter
            
            # For length measurements, use appropriate planar datums
            elif any(keyword in fname for keyword in ["plane", "length", "distance"]):
                # Look for datum A first (often base plane)
                if 'A' in datum_results:
                    return 'A'
                # Then look for any planar datum
                for datum_letter, location in datum_results.items():
                    if any(keyword in location.lower() for keyword in ["face", "plane"]):
                        return datum_letter
            
            # If all else fails, return the first available datum
            return list(datum_results.keys())[0]
        
        return ""

    def extract_dimensional_tolerances(self, text, line_dict, datum_results=None, datum_letter_to_feature=None):
        """ENHANCED: Extract dimensional tolerances (diameter, length) with their tolerance values and datums"""
        dimensional_results = []
        
        # Ensure we have datum information
        if datum_results is None:
            datum_results = {}
        if datum_letter_to_feature is None:
            datum_letter_to_feature = {}
        
        try:
            # Pattern to find DIMENSIONAL_SIZE for diameter
            dimensional_size_pattern = re.compile(
                r"(#\d+)\s*=\s*DIMENSIONAL_SIZE\s*\(\s*(#\d+)\s*,\s*'([^']*)'\s*\)", re.IGNORECASE
            )
            
            # Pattern to find DIMENSIONAL_LOCATION for length/distance
            dimensional_location_pattern = re.compile(
                r"(#\d+)\s*=\s*DIMENSIONAL_LOCATION\s*\(\s*'([^']*)'[^#]*#(\d+)[^#]*#(\d+)", re.IGNORECASE
            )
            
            # Pattern to find PLUS_MINUS_TOLERANCE
            plus_minus_tolerance_pattern = re.compile(
                r"(#\d+)\s*=\s*PLUS_MINUS_TOLERANCE\s*\(\s*(#\d+)\s*,\s*(#\d+)\s*\)", re.IGNORECASE
            )
            
            # Pattern to find TOLERANCE_VALUE with upper and lower limits
            tolerance_value_pattern = re.compile(
                r"(#\d+)\s*=\s*TOLERANCE_VALUE\s*\(\s*(#\d+)\s*,\s*(#\d+)\s*\)", re.IGNORECASE
            )
            
            # Pattern to find LENGTH_MEASURE_WITH_UNIT for nominal values and tolerances
            length_measure_pattern = re.compile(
                r"(#\d+)\s*=.*?LENGTH_MEASURE\(([^)]+)\)", re.IGNORECASE
            )
            
            # Pattern to find nominal values in MEASURE_WITH_UNIT
            nominal_value_pattern = re.compile(
                r"POSITIVE_LENGTH_MEASURE\(([^)]+)\)", re.IGNORECASE
            )
            
            # Build mapping of tolerance relationships
            tolerance_map = {}
            dimension_to_tolerance = {}
            tolerance_values = {}
            length_measures = {}
            nominal_values = {}
            
            # Extract LENGTH_MEASURE values
            for match in length_measure_pattern.finditer(text):
                entity_id, measure_value = match.groups()
                try:
                    length_measures[entity_id] = float(measure_value)
                except ValueError:
                    pass
            
            # Extract nominal values
            for line in text.splitlines():
                nominal_match = nominal_value_pattern.search(line)
                if nominal_match:
                    entity_match = re.match(r"(#\d+)\s*=", line)
                    if entity_match:
                        try:
                            nominal_values[entity_match.group(1)] = float(nominal_match.group(1))
                        except ValueError:
                            pass
            
            # Extract TOLERANCE_VALUE relationships (upper and lower limits)
            for match in tolerance_value_pattern.finditer(text):
                tol_val_id, lower_ref, upper_ref = match.groups()
                lower_value = length_measures.get(lower_ref, 0)
                upper_value = length_measures.get(upper_ref, 0)
                tolerance_values[tol_val_id] = {
                    'lower': lower_value,
                    'upper': upper_value,
                    'range': abs(upper_value - lower_value)
                }
            
            # Extract PLUS_MINUS_TOLERANCE relationships
            for match in plus_minus_tolerance_pattern.finditer(text):
                pm_tol_id, tolerance_val_ref, dimension_ref = match.groups()
                dimension_to_tolerance[dimension_ref] = {
                    'tolerance_id': tolerance_val_ref,
                    'pm_tolerance_id': pm_tol_id
                }
            
            # Process DIMENSIONAL_SIZE (diameters)
            for match in dimensional_size_pattern.finditer(text):
                dim_size_id, shape_aspect_ref, dim_type = match.groups()
                
                # Get the shape aspect information
                shape_aspect_line = line_dict.get(shape_aspect_ref, "")
                feature_name = ""
                diameter_value = ""
                
                # Extract feature name from shape aspect
                shape_aspect_match = re.search(r"SHAPE_ASPECT\s*\(\s*'([^']*)'", shape_aspect_line)
                if shape_aspect_match:
                    feature_name = shape_aspect_match.group(1)
                
                # Find associated nominal value
                for line in text.splitlines():
                    if dim_size_id in line and "DIMENSIONAL_CHARACTERISTIC_REPRESENTATION" in line:
                        # Find the representation reference
                        repr_match = re.search(r"#(\d+)\)", line)
                        if repr_match:
                            repr_line = line_dict.get(f"#{repr_match.group(1)}", "")
                            # Find nominal value reference in representation
                            nominal_ref_match = re.search(r"\(#(\d+)\)", repr_line)
                            if nominal_ref_match:
                                nominal_ref = f"#{nominal_ref_match.group(1)}"
                                diameter_value = nominal_values.get(nominal_ref, "")
                
                # Get tolerance information
                tolerance_info = dimension_to_tolerance.get(dim_size_id, {})
                tolerance_range = ""
                upper_limit = ""
                lower_limit = ""
                
                if tolerance_info:
                    tol_val_id = tolerance_info.get('tolerance_id')
                    if tol_val_id in tolerance_values:
                        tol_data = tolerance_values[tol_val_id]
                        tolerance_range = f"±{tol_data['range']/2:.3f}"
                        if diameter_value:
                            try:
                                nom_val = float(diameter_value)
                                upper_limit = f"{nom_val + tol_data['upper']:.3f}"
                                lower_limit = f"{nom_val + tol_data['lower']:.3f}"
                            except ValueError:
                                pass
                
                # ENHANCED: Find appropriate datum for this dimensional tolerance
                datum_letter = self.find_datum_for_dimensional_tolerance(
                    feature_name, datum_results, datum_letter_to_feature
                )
                
                # Determine location based on feature name
                location = "cylindrical surface"
                surface = "curved side of the cylinder"
                if "boss" in feature_name.lower():
                    location = "cylindrical side"
                    surface = "curved side of the cylinder"
                
                # Create symbol with type
                symbol = self.gdnt_symbols.get("Diameter", "⌀")
                type_with_symbol = f"{symbol} Diameter"
                
                dimensional_results.append((
                    type_with_symbol, 
                    str(diameter_value) if diameter_value else "N/A",
                    datum_letter,  # NOW INCLUDES DATUM
                    location,
                    surface,
                    tolerance_range,
                    upper_limit,
                    lower_limit
                ))
            
            # Process DIMENSIONAL_LOCATION (lengths/distances)
            for match in dimensional_location_pattern.finditer(text):
                dim_loc_id, distance_type, shape_aspect1_id, shape_aspect2_id = match.groups()
                
                # Get shape aspect information for both references
                shape_aspect1_line = line_dict.get(f"#{shape_aspect1_id}", "")
                shape_aspect2_line = line_dict.get(f"#{shape_aspect2_id}", "")
                
                feature_name1 = ""
                feature_name2 = ""
                
                shape_match1 = re.search(r"SHAPE_ASPECT\s*\(\s*'([^']*)'", shape_aspect1_line)
                shape_match2 = re.search(r"SHAPE_ASPECT\s*\(\s*'([^']*)'", shape_aspect2_line)
                
                if shape_match1:
                    feature_name1 = shape_match1.group(1)
                if shape_match2:
                    feature_name2 = shape_match2.group(1)
                
                # Combine feature names for datum detection
                combined_feature_name = f"{feature_name1} to {feature_name2}"
                
                # Find nominal distance value
                distance_value = ""
                for line in text.splitlines():
                    if dim_loc_id in line and "DIMENSIONAL_CHARACTERISTIC_REPRESENTATION" in line:
                        repr_match = re.search(r"#(\d+)\)", line)
                        if repr_match:
                            repr_line = line_dict.get(f"#{repr_match.group(1)}", "")
                            nominal_ref_match = re.search(r"\(#(\d+)\)", repr_line)
                            if nominal_ref_match:
                                nominal_ref = f"#{nominal_ref_match.group(1)}"
                                distance_value = nominal_values.get(nominal_ref, "")
                
                # Get tolerance information
                tolerance_info = dimension_to_tolerance.get(dim_loc_id, {})
                tolerance_range = ""
                upper_limit = ""
                lower_limit = ""
                
                if tolerance_info:
                    tol_val_id = tolerance_info.get('tolerance_id')
                    if tol_val_id in tolerance_values:
                        tol_data = tolerance_values[tol_val_id]
                        tolerance_range = f"±{tol_data['range']/2:.3f}"
                        if distance_value:
                            try:
                                nom_val = float(distance_value)
                                upper_limit = f"{nom_val + tol_data['upper']:.3f}"
                                lower_limit = f"{nom_val + tol_data['lower']:.3f}"
                            except ValueError:
                                pass
                
                # ENHANCED: Find appropriate datum for this dimensional tolerance
                # For length measurements, try both feature names
                datum_letter = ""
                if feature_name1:
                    datum_letter = self.find_datum_for_dimensional_tolerance(
                        feature_name1, datum_results, datum_letter_to_feature
                    )
                if not datum_letter and feature_name2:
                    datum_letter = self.find_datum_for_dimensional_tolerance(
                        feature_name2, datum_results, datum_letter_to_feature
                    )
                
                # If still no datum found, use the combined name
                if not datum_letter:
                    datum_letter = self.find_datum_for_dimensional_tolerance(
                        combined_feature_name, datum_results, datum_letter_to_feature
                    )
                
                # Determine location based on feature names
                location = "between surfaces"
                surface = "linear distance"
                
                if feature_name1 and feature_name2:
                    if "plane" in feature_name1.lower() and "plane" in feature_name2.lower():
                        location = "between planes"
                        surface = "planar faces"
                
                # Create symbol with type
                symbol = self.gdnt_symbols.get("Linear Distance", "↔")
                type_with_symbol = f"{symbol} {distance_type.title()}" if distance_type else f"{symbol} Length"
                
                dimensional_results.append((
                    type_with_symbol,
                    str(distance_value) if distance_value else "N/A",
                    datum_letter,  # NOW INCLUDES DATUM
                    location,
                    surface,
                    tolerance_range,
                    upper_limit,
                    lower_limit
                ))
            
            return dimensional_results
            
        except Exception as e:
            logger.error(f"Error extracting dimensional tolerances: {str(e)}")
            return []

    def extract_tolerance_data(self, text):
        """Extract tolerance values and datums from STEP/text file with enhanced datum detection"""
        try:
            lines = text.splitlines()
            
            # Build entity mapping
            line_dict = {}
            for line in lines:
                match = re.match(r"(#\d+)\s*=", line)
                if match:
                    line_dict[match.group(1)] = line.strip()

            # Enhanced regex patterns for tolerance, datum, and shape aspect entities
            tol_pattern = re.compile(
                r"(#\d+)\s*=\s*(CYLINDRICITY|FLATNESS|STRAIGHTNESS|ROUNDNESS)_TOLERANCE"
                r"\(\s*'([^']*)'\s*,\s*''\s*,\s*(#\d+)", re.IGNORECASE
            )
            
            # Updated datum pattern to handle the new format
            datum_pattern = re.compile(
                r"#\d+\s*=\s*DATUM\('([^']*)',\$,#\d+,\.F\.,'([A-Z])'\);", re.IGNORECASE
            )
            
            shape_aspect_pattern = re.compile(
                r"#\d+\s*=\s*SHAPE_ASPECT\('([^']*?)\((\w)?'?,.*?#(\d+)\)"
            )
            
            # Data structures
            tol_results = []
            datum_results = {}
            face_to_plane = {}
            
            # Enhanced datum extraction - Build datum letter to feature name mapping
            datum_letter_to_feature = {}
            datum_id_to_letter = {}  # Map entity IDs to datum letters
            
            # Find all DATUM lines with the enhanced pattern
            for line in lines:
                # Updated regex to match your file format: #522=DATUM('Datum29@Boss1(A)',$,#23,.F.,'A');
                datum_match = re.match(
                    r"#(\d+)=DATUM\('([^']*)',\$,#\d+,\.F\.,'([A-Z])'\);", line)
                if datum_match:
                    datum_id, feature_name, datum_letter = datum_match.groups()
                    datum_letter_to_feature[datum_letter] = feature_name
                    datum_id_to_letter[datum_id] = datum_letter
                    
                    # Extract the geometric feature type from the feature name
                    # Example: 'Datum29@Boss1(A)' -> 'Boss1', 'Datum28@Plane1(D)' -> 'Plane1'
                    feature_match = re.search(r'@([^(]+)', feature_name)
                    if feature_match:
                        geometric_feature = feature_match.group(1).lower()
                        if 'boss' in geometric_feature:
                            datum_results[datum_letter] = "cylindrical side"
                        elif 'plane1' in geometric_feature:
                            datum_results[datum_letter] = "bottom face"
                        elif 'plane2' in geometric_feature:
                            datum_results[datum_letter] = "top face"
                        elif 'plane' in geometric_feature:
                            datum_results[datum_letter] = self.determine_plane_position(geometric_feature, datum_letter)
                        else:
                            datum_results[datum_letter] = geometric_feature
                    else:
                        # Fallback if no @ pattern found
                        datum_results[datum_letter] = self.determine_plane_position(feature_name, datum_letter)

            # Map face IDs to shape descriptions (Plane1, Plane2, Boss1, etc.)
            for match in shape_aspect_pattern.finditer(text):
                shape_name, datum_letter, plane_id = match.groups()
                shape_name = shape_name.lower()
                location = ""
                if "plane1" in shape_name:
                    location = "Plane1"
                elif "plane2" in shape_name:
                    location = "Plane2"
                elif "boss1" in shape_name:
                    location = "Boss1"
                elif "torus" in shape_name:
                    location = "torus side"
                elif "top" in shape_name:
                    location = "top face"
                elif "bottom" in shape_name:
                    location = "bottom face"
                elif "cylindrical" in shape_name or "side" in shape_name:
                    location = "cylindrical side"
                face_to_plane[plane_id] = location
                if datum_letter:
                    datum_results[datum_letter] = location

            # Extract geometric tolerances and associate with datums and locations
            for tol_id, tol_type, tol_name, ref_id in tol_pattern.findall(text):
                definition = line_dict.get(ref_id, "")
                value_match = re.search(
                    r"(?:LENGTH_MEASURE|VALUE_REPRESENTATION_ITEM)\s*\(\s*([\d.]+)", 
                    definition
                )
                value = value_match.group(1) if value_match else "N/A"
                label = "Circularity" if tol_type.upper() == "ROUNDNESS" else tol_type.capitalize()

                # Enhanced datum letter and location detection
                datum_letter = ""
                location = ""
                
                # Method 1: Extract datum letter from tolerance name (e.g., "tolerance(A)")
                tol_name_lower = tol_name.lower()
                datum_match = re.search(r'\(([A-Z])\)', tol_name)
                if datum_match:
                    datum_letter = datum_match.group(1)
                    location = datum_results.get(datum_letter, "")
                
                # Method 2: Try to find datum letter from tolerance name with lowercase
                if not datum_letter:
                    for d_letter in datum_results:
                        if f"({d_letter.lower()})" in tol_name_lower or f"({d_letter.upper()})" in tol_name:
                            datum_letter = d_letter
                            location = datum_results[d_letter]
                            break
                
                # Method 3: Look for DATUM_FEATURE references in the tolerance line
                if not datum_letter:
                    tol_line = line_dict.get(tol_id, "")
                    # Find all entity references in the tolerance line
                    entity_refs = re.findall(r'#(\d+)', tol_line)
                    
                    # Check each reference to see if it leads to a datum
                    for entity_ref in entity_refs:
                        # Look for DATUM_FEATURE that references this entity
                        for line in lines:
                            if f"DATUM_FEATURE" in line and f"#{entity_ref}" in line:
                                # Extract datum letter from DATUM_FEATURE line
                                df_match = re.search(r"DATUM_FEATURE\([^']*'[^']*\(([A-Z])\)'", line)
                                if df_match:
                                    datum_letter = df_match.group(1)
                                    location = datum_results.get(datum_letter, "")
                                    break
                        if datum_letter:
                            break
                
                # Method 4: Check if tolerance name contains feature references
                if not location:
                    # Look for Boss1, Plane1, Plane2 in tolerance name
                    if "boss" in tol_name_lower:
                        location = "cylindrical side"
                        # Try to find corresponding datum
                        for d_letter, d_feature in datum_letter_to_feature.items():
                            if "boss" in d_feature.lower() and not datum_letter:
                                datum_letter = d_letter
                                break
                    elif "plane1" in tol_name_lower:
                        location = "bottom face"
                        # Try to find corresponding datum
                        for d_letter, d_feature in datum_letter_to_feature.items():
                            if "plane1" in d_feature.lower() and not datum_letter:
                                datum_letter = d_letter
                                break
                    elif "plane2" in tol_name_lower:
                        location = "top face"
                        # Try to find corresponding datum
                        for d_letter, d_feature in datum_letter_to_feature.items():
                            if "plane2" in d_feature.lower() and not datum_letter:
                                datum_letter = d_letter
                                break
                    elif "plane" in tol_name_lower:
                        location = "planar surface"
                
                # Method 5: Smart matching based on tolerance type and surface type
                if not datum_letter and location:
                    for d_letter in datum_results:
                        d_location = datum_results[d_letter]
                        # Match based on surface compatibility
                        if ((location == "cylindrical side" and d_location == "cylindrical side") or
                            (location in ["bottom face", "top face", "planar surface"] and 
                             d_location in ["bottom face", "top face", "planar surface"])):
                            datum_letter = d_letter
                            break
                
                # Method 6: Infer location based on tolerance type and find matching datum
                if not location:
                    if label.lower() in ["cylindricity", "circularity"]:
                        location = "cylindrical side"
                        # Find cylindrical datum
                        for d_letter in datum_results:
                            if datum_results[d_letter] == "cylindrical side":
                                datum_letter = d_letter
                                break
                    elif label.lower() in ["flatness"]:
                        location = "planar surface"
                        # Find planar datum
                        for d_letter in datum_results:
                            if "face" in datum_results[d_letter]:
                                datum_letter = d_letter
                                location = datum_results[d_letter]
                                break
                    elif label.lower() in ["straightness"]:
                        # For straightness, determine location based on available datums and context
                        # Straightness can be applied to various surfaces
                        available_datums = list(datum_results.keys())
                        if available_datums:
                            # Use a smarter approach: match with context or sequence
                            current_count = len(tol_results)
                            if current_count < len(available_datums):
                                datum_letter = available_datums[current_count % len(available_datums)]
                                location = datum_results[datum_letter]
                            else:
                                datum_letter = available_datums[0]
                                location = datum_results[datum_letter]

                # Keep tolerance value, upper and lower limits empty as requested for geometric tolerances
                tolerance_value = ""
                upper_limit = ""
                lower_limit = ""

                tol_results.append((label, value, datum_letter, location, tolerance_value, upper_limit, lower_limit))

            # ENHANCED: Extract dimensional tolerances with datum information
            dimensional_tolerances = self.extract_dimensional_tolerances(
                text, line_dict, datum_results, datum_letter_to_feature
            )
            
            # Build table rows for results
            table_rows = []
            
            # Add geometric tolerances
            for label, value, datum, loc, tolerance_value, upper_limit, lower_limit in tol_results:
                symbol = self.gdnt_symbols.get(label, "")
                type_with_symbol = f"{symbol} {label}" if symbol else label
                
                # MODIFIED: Use the actual surface location directly instead of "at datum X"
                location_str = loc if loc else "surface"
                
                # Map location to surface description for consistency
                if loc == "cylindrical side":
                    surface = "curved side of the cylinder"
                elif loc == "bottom face":
                    surface = "bottom face"
                elif loc == "top face":
                    surface = "top face"  
                elif loc == "planar surface":
                    surface = "planar face"
                else:
                    surface = loc if loc else "surface"
                    
                table_rows.append((type_with_symbol, value, datum, location_str, surface, tolerance_value, upper_limit, lower_limit))
            
            # Add dimensional tolerances (ENHANCED with datums)
            for dim_tolerance in dimensional_tolerances:
                table_rows.append(dim_tolerance)
            
            # Enhanced datum processing with better plane detection and feature extraction
            for d_letter, feature_name in datum_letter_to_feature.items():
                location_str = datum_results.get(d_letter, "surface")
                surface = location_str
                
                # For datums, keep all three new columns empty
                table_rows.append(("Datum", d_letter, d_letter, location_str, surface, "", "", ""))
                
            return table_rows
            
        except Exception as e:
            logger.error(f"Error extracting tolerance data: {str(e)}")
            return []

    def create_interface(self):
        """Create the main GUI interface"""
        self.root = tk.Tk()
        self.root.title("GD&T Tolerance Value Extractor  -  Dimensional Tolerances")
        self.root.geometry("1500x750")
        self.root.configure(bg="#eef3f7")
        
        # Create frames
        self.create_header()
        self.create_control_panel()
        self.create_table()
        self.create_status_bar()
        
        # Event bindings
        self.root.bind('<Control-o>', lambda e: self.upload_and_process())
        self.root.bind('<Control-s>', lambda e: self.save_results())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        
        self.root.mainloop()

    def create_header(self):
        """Create header with logos and title - centered layout"""
        self.top_frame = tk.Frame(self.root, bg="#eef3f7")
        self.top_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Configure grid weights for centering
        self.top_frame.grid_columnconfigure(0, weight=1)  # Left spacer
        self.top_frame.grid_columnconfigure(1, weight=0)  # Left logo
        self.top_frame.grid_columnconfigure(2, weight=0)  # Title area
        self.top_frame.grid_columnconfigure(3, weight=0)  # Right logo
        self.top_frame.grid_columnconfigure(4, weight=1)  # Right spacer
        
        # Left logo
        logo_frame = tk.Frame(self.top_frame, bg="#ffffff", borderwidth=1, relief="solid")
        logo_frame.grid(row=0, column=1, sticky="", padx=10, rowspan=2)
        
        try:
            logo_img = Image.open("./assets/unisza1.png").resize((100, 100))
            self.logo_photo = ImageTk.PhotoImage(logo_img)
            tk.Label(logo_frame, image=self.logo_photo, bg="#ffffff").pack()
        except FileNotFoundError:
            tk.Label(logo_frame, text="LOGO\n1", bg="#ffffff", 
                    width=12, height=6, font=("Arial", 8)).pack()

        # Title and subtitle frame - centered
        title_frame = tk.Frame(self.top_frame, bg="#eef3f7")
        title_frame.grid(row=0, column=2, rowspan=2, padx=30, pady=10)
        
        # Title
        self.title_label = tk.Label(
            title_frame, 
            text="GD&T Tolerance Value Extractor",
            font=("Helvetica", 20, "bold"), 
            bg="#eef3f7", 
            fg="#003366"
        )
        self.title_label.pack(pady=(0, 5))
        
        # Subtitle
        self.subtitle_label = tk.Label(
            title_frame, 
            text="Enhanced with Dimensional Tolerance Extraction",
            font=("Helvetica", 12, "italic"), 
            bg="#eef3f7", 
            fg="#666666"
        )
        self.subtitle_label.pack()

        # Right logo
        right_logo_frame = tk.Frame(self.top_frame, bg="#ffffff", borderwidth=1, relief="solid")
        right_logo_frame.grid(row=0, column=3, sticky="", padx=10, rowspan=2)
        
        try:
            right_img = Image.open("./assets/frit.png").resize((100, 100))
            self.right_photo = ImageTk.PhotoImage(right_img)
            tk.Label(right_logo_frame, image=self.right_photo, bg="#ffffff").pack()
        except FileNotFoundError:
            tk.Label(right_logo_frame, text="LOGO\n2", bg="#ffffff", 
                    width=12, height=6, font=("Arial", 8)).pack()

    def create_control_panel(self):
        """Create control buttons and options"""
        self.mid_frame = tk.Frame(self.root, bg="#eef3f7")
        self.mid_frame.pack(pady=5)

        # Buttons
        tk.Button(
            self.mid_frame, 
            text="📤 Upload STEP File (Ctrl+O)", 
            command=self.upload_and_process,
            font=("Arial", 11, "bold"), 
            bg="#4CAF50", 
            fg="white",
            relief="raised",
            bd=2
        ).grid(row=0, column=0, padx=5, pady=2)

        tk.Button(
            self.mid_frame, 
            text="💾 Save Results (Ctrl+S)", 
            command=self.save_results,
            font=("Arial", 11), 
            bg="#2196F3", 
            fg="white",
            relief="raised",
            bd=2
        ).grid(row=0, column=1, padx=5, pady=2)

        tk.Button(
            self.mid_frame, 
            text="🗑️ Clear Output", 
            command=self.clear_output,
            font=("Arial", 11), 
            bg="#FF9800", 
            fg="white",
            relief="raised",
            bd=2
        ).grid(row=0, column=2, padx=5, pady=2)

        self.theme_button = tk.Button(
            self.mid_frame, 
            text="🌙 Dark Mode", 
            command=self.toggle_theme,
            font=("Arial", 10), 
            bg="#9E9E9E", 
            fg="black",
            relief="raised",
            bd=2
        )
        self.theme_button.grid(row=0, column=3, padx=5, pady=2)

        # Format selection
        tk.Label(self.mid_frame, text="Export:", bg="#eef3f7", 
                font=("Arial", 10)).grid(row=0, column=4, padx=(10,5), pady=2)
        
        self.format_var = tk.StringVar(value=".xlsx")
        format_menu = tk.OptionMenu(self.mid_frame, self.format_var, ".txt", ".csv", ".xlsx")
        format_menu.config(font=("Arial", 10), bg="#f0f0f0")
        format_menu.grid(row=0, column=5, padx=5, pady=2)

        # Processing info
        info_label = tk.Label(
            self.mid_frame, 
            text="✨ Extracts: Geometric Tolerances + Dimensional Tolerances (Diameter & Length) + Datums", 
            font=("Arial", 9, "italic"), 
            bg="#eef3f7", 
            fg="#008000"
        )
        info_label.grid(row=1, column=0, columnspan=6, pady=(5,0))

        # File status
        self.filename_label = tk.Label(
            self.root, 
            text="No file loaded", 
            font=("Arial", 10), 
            bg="#eef3f7", 
            fg="#666666"
        )
        self.filename_label.pack(pady=(5, 0))

    def create_table(self):
        """Create the results table"""
        # Table frame with scrollbars
        table_frame = tk.Frame(self.root)
        table_frame.pack(expand=True, fill='both', padx=20, pady=10)

        # Treeview with scrollbars - Updated columns to include the new three columns
        columns = ("Type", "Value", "Datum", "Location", "Surface", "Tolerance Value", "Upper Limit", "Lower Limit")
        self.output_table = ttk.Treeview(
            table_frame, 
            columns=columns, 
            show="headings", 
            height=22
        )

        # Configure columns with updated widths
        column_widths = {
            "Type": 140, 
            "Value": 80, 
            "Datum": 60, 
            "Location": 140, 
            "Surface": 140, 
            "Tolerance Value": 120,
            "Upper Limit": 100, 
            "Lower Limit": 100
        }
        for col in columns:
            self.output_table.heading(col, text=col, command=lambda c=col: self.sort_table(c))
            self.output_table.column(col, anchor="center", width=column_widths.get(col, 100))

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.output_table.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.output_table.xview)
        
        self.output_table.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid layout
        self.output_table.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Configure row colors
        self.output_table.tag_configure('even', background='#f8f9fa')
        self.output_table.tag_configure('odd', background='#e9ecef')
        self.output_table.tag_configure('datum', background='#fff3cd', foreground='#856404')
        self.output_table.tag_configure('dimensional', background='#e8f5e8', foreground='#2e7d32')

    def create_status_bar(self):
        """Create status bar"""
        self.status_bar = tk.Label(
            self.root, 
            text="Ready - Enhanced to extract dimensional tolerances with datums", 
            relief=tk.SUNKEN, 
            anchor=tk.W,
            bg="#e9ecef",
            font=("Arial", 9)
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def upload_and_process(self):
        """Handle file upload and processing"""
        file_path = filedialog.askopenfilename(
            title="Select STEP File",
            filetypes=[
                ("STEP files", "*.step *.stp"),
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
            
        self.status_bar.config(text="Processing file... (extracting geometric and dimensional tolerances with datums)")
        self.root.update()
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
                
            result_data = self.extract_tolerance_data(content)
            self.show_table(result_data)
            
            # Count different types of extractions
            geometric_count = len([r for r in result_data if not any(symbol in r[0] for symbol in ["⌀ Diameter", "↔"]) and r[0] != "Datum"])
            dimensional_count = len([r for r in result_data if any(symbol in r[0] for symbol in ["⌀ Diameter", "↔"])])
            datum_count = len([r for r in result_data if r[0] == "Datum"])
            
            filename = os.path.basename(file_path)
            self.filename_label.config(text=f"📁 Loaded: {filename}")
            self.status_bar.config(text=f"✅ Extracted: {geometric_count} geometric, {dimensional_count} dimensional (with datums), {datum_count} datums from {filename}")
            
        except Exception as e:
            error_msg = f"Error reading file: {str(e)}"
            messagebox.showerror("File Error", error_msg)
            logger.error(error_msg)
            self.status_bar.config(text="❌ Error processing file")

    def show_table(self, result_lines):
        """Display results in the table"""
        # Clear existing data
        for row in self.output_table.get_children():
            self.output_table.delete(row)

        # Insert new data with enhanced tagging
        for idx, line in enumerate(result_lines):
            if line[0] == 'Datum':
                tag = 'datum'
            elif any(symbol in line[0] for symbol in ["⌀ Diameter", "↔"]):
                tag = 'dimensional'
            else:
                tag = 'even' if idx % 2 == 0 else 'odd'
            
            self.output_table.insert("", "end", values=line, tags=(tag,))

    def sort_table(self, column):
        """Sort table by column"""
        data = [(self.output_table.set(child, column), child) 
                for child in self.output_table.get_children()]
        data.sort(key=lambda x: x[0])
        
        for index, (val, child) in enumerate(data):
            self.output_table.move(child, '', index)

    def save_results(self):
        """Save results to file"""
        if not self.output_table.get_children():
            messagebox.showwarning("Warning", "No data to save.")
            return

        rows = [self.output_table.item(row)['values'] for row in self.output_table.get_children()]
        file_format = self.format_var.get()
        
        filetypes = {
            ".txt": [("Text files", "*.txt")],
            ".csv": [("CSV files", "*.csv")], 
            ".xlsx": [("Excel files", "*.xlsx")]
        }
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=file_format,
            filetypes=filetypes[file_format],
            title=f"Save as {file_format.upper()}"
        )
        
        if not file_path:
            return

        try:
            headers = ["Type", "Value", "Datum", "Location", "Surface", "Tolerance Value", "Upper Limit", "Lower Limit"]
            
            if file_format == ".txt":
                with open(file_path, "w", encoding="utf-8") as file:
                    file.write("\t".join(headers) + "\n")
                    file.write("="*120 + "\n")
                    for row in rows:
                        file.write("\t".join(str(cell) for cell in row) + "\n")
                        
            elif file_format == ".csv":
                import csv
                with open(file_path, "w", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(headers)
                    writer.writerows(rows)
                    
            elif file_format == ".xlsx":
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "GD&T & Dimensional Tolerances"
                
                # Define colors
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                datum_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                dimensional_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                tolerance_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                
                # Add headers with formatting
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data with enhanced formatting
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Apply conditional formatting
                        if row[0] == "Datum":
                            cell.fill = datum_fill
                            if col_idx == 1:  # Type column
                                cell.font = Font(bold=True)
                        elif any(symbol in row[0] for symbol in ["⌀ Diameter", "↔"]):
                            cell.fill = dimensional_fill
                            if col_idx == 1:  # Type column
                                cell.font = Font(bold=True, color="2E7D32")
                        elif col_idx in [6, 7, 8]:  # Tolerance Value, Upper Limit, Lower Limit columns
                            if value and value != "N/A" and value != "":
                                cell.fill = tolerance_fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
                
                # Add summary
                last_row = len(rows) + 3
                ws.cell(row=last_row, column=1, value="SUMMARY:").font = Font(bold=True)
                
                geometric_count = len([r for r in rows if not any(symbol in r[0] for symbol in ["⌀ Diameter", "↔"]) and r[0] != "Datum"])
                dimensional_count = len([r for r in rows if any(symbol in r[0] for symbol in ["⌀ Diameter", "↔"])])
                datum_count = len([r for r in rows if r[0] == "Datum"])
                dimensional_with_datum_count = len([r for r in rows if any(symbol in r[0] for symbol in ["⌀ Diameter", "↔"]) and r[2] != ""])
                
                ws.cell(row=last_row+1, column=1, value=f"Geometric Tolerances: {geometric_count}")
                ws.cell(row=last_row+2, column=1, value=f"Dimensional Tolerances: {dimensional_count}")
                ws.cell(row=last_row+3, column=1, value=f"Dimensional with Datums: {dimensional_with_datum_count}")
                ws.cell(row=last_row+4, column=1, value=f"Datums: {datum_count}")
                ws.cell(row=last_row+5, column=1, value=f"Total Items: {len(rows)}")
                
                wb.save(file_path)
                
            self.status_bar.config(text=f"✅ Results saved to {file_format.upper()} format")
            messagebox.showinfo("Success", f"Results saved as {file_format.upper()}!\n\nIncludes geometric and dimensional tolerances with datum references.")
            
        except Exception as e:
            error_msg = f"Failed to save file: {str(e)}"
            messagebox.showerror("Save Error", error_msg)
            logger.error(error_msg)

    def clear_output(self):
        """Clear output and reset interface"""
        if self.output_table:
            for row in self.output_table.get_children():
                self.output_table.delete(row)
        self.filename_label.config(text="No file loaded")
        self.status_bar.config(text="Ready - Enhanced to extract dimensional tolerances with datums")

    def toggle_theme(self):
        """Toggle between dark and light themes"""
        self.dark_mode = not self.dark_mode
        
        # Color schemes
        if self.dark_mode:
            bg_color, fg_color = "#2c2f33", "#f0f0f0"
            button_bg, button_fg = "#7289da", "white"
            theme_text = "☀️ Light Mode"
        else:
            bg_color, fg_color = "#eef3f7", "#003366"
            button_bg, button_fg = "#9E9E9E", "black" 
            theme_text = "🌙 Dark Mode"

        # Apply theme to title frame
        widgets_to_update = [
            (self.root, {"bg": bg_color}),
            (self.top_frame, {"bg": bg_color}),
            (self.mid_frame, {"bg": bg_color}),
            (self.title_label, {"bg": bg_color, "fg": fg_color}),
            (self.subtitle_label, {"bg": bg_color, "fg": "#bbbbbb" if self.dark_mode else "#666666"}),
            (self.filename_label, {"bg": bg_color, "fg": "#bbbbbb" if self.dark_mode else "black"}),
            (self.theme_button, {"text": theme_text, "bg": button_bg, "fg": button_fg}),
            (self.status_bar, {"bg": "#1e1e1e" if self.dark_mode else "#e9ecef", 
                              "fg": fg_color})
        ]
        
        # Update title frame background if it exists
        try:
            title_frame = self.title_label.master
            title_frame.configure(bg=bg_color)
        except:
            pass
        
        for widget, config in widgets_to_update:
            widget.configure(**config)

def main():
    """Main application entry point"""
    try:
        app = GDTExtractor()
        app.create_interface()
    except Exception as e:
        logging.error(f"Application failed to start: {str(e)}")
        messagebox.showerror("Startup Error", f"Failed to start application:\n{str(e)}")

if __name__ == "__main__":
    main()